"""
SINOBY MediaPlanBot
"""

import asyncio
import base64
import io
import json
import logging
import os
import random
import tempfile
from datetime import date
from calendar import monthrange
from urllib.parse import urlencode

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message, WebAppInfo, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
    CallbackQuery, BufferedInputFile,
)
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image as PILImage

load_dotenv()

BOT_TOKEN  = os.getenv("BOT_TOKEN")
WEBAPP_URL = os.getenv("WEBAPP_URL", "https://drouchucha.github.io/sinoby-mediaplan/")
SHEETS_URL = "https://script.google.com/macros/s/AKfycbwORRGftN-ehlSFpJlvNk74IxoDLlGQ8BW_weDDztzK7bQSE3z29QeBg5beTTP_tNrF8A/exec"
MANAGER_IDS = [267728315]

# Логотип встроен как base64
LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAlgAAACSCAIAAAAW6CNwAABYbUlEQVR42u19d1hU19b3PmUKVTqIioqgQcWGDdFg77HRRLCXaOwlaopGE2M0atRobCFGjQ0VjA2NvWCLBWPBDooFkSp9Zk75/ljX883LnBnKDMOM7t9znzxeZuacXVZfa69N8DyPMDAwMDAwPlaQeAkwMDAwMLAixMDAwMDAwIoQAwMDAwMDK0IMDAwMDAysCDEwMDAwMD4W0CY1GihhVS9kJQhC+C8GBkalsh7mPoyPE0QVHp/geZ7jOBgASZIkqcs95TiO4zhgS5IkMXNiYOjDegJK5T51PsXch4EVoSGVH0VRmuxUWFhYWFhYXFzMMAxCiKZpuVxuaWlpaWmp+RyWZTFbYmBUgPtomtY0NAsKCoqKioqLi4GzJBKJpaWlhYWFXC4XtUpLNV4xMLAiFHfp1Dnw7du3iYmJd+7cuX//flJS0ps3bzIzMwsKCpRKJcuywK5SqdTS0tLR0dHNza1u3bo+Pj6NGzdu2LBh9erVhecwDEMQBEVRVSVeEFLXxBzP84iAv8MHJOJJRJT4mqlDcMF1Q9Sg0QHYWR1f0JTRZqFgWJbVxWZVSZ/i+u/58+d37969e/fugwcPkpOT09PTs7KyioqKVCoVKEKapmUymbW1tZOTU40aNby8vBo1atS4ceMGDRpUq1ZNfeJVrhHBbv6QxbSB6Ef3QlUtlVYG3wHKQp+VrgiBCQW/rbi4+Nq1aydPnjx37tzt27ezs7Mr8Mxq1ao1atQoMDCwW7durVu3trKyEoRsFTmIPEIsjziECAJJRJQKzxAEIhBtRhSG/ewPwwVkWVbQf5mZmfHx8SdOnIiPj3/06FFRUVEFnlm9evUWLVp06tSpS5cuzZo1E8wmCPPgNcfAHqGIChR449q1a9HR0YcOHXr06JGmrhbSFZrjETL28I8SnkrdunX79OkTFhbWvn37qlKHPFISvBQRCPFIxWcWcy9YrognEIFIKWFvSdVDBIV4xBMcgQjT9wvBcHny5Mnp06dJktTmF5IkybJsly5dvLy84CdlcQf37t2bm5tLECUJD/4il8tDQ0M1Y3EmbjG8efPm4MGDmpMSVsnLy6tLly7GNC/UuU+hUJw8eTI6OvrEiRNv3rzR9ObV84W6uU/doSdJ0s/Pb+DAgcHBwd7e3lWlDhUKxZ49e4qKikTX/wPwBXmet7KyCgsL0zNYolKpdu/eXVhYSJKkKPdZW1uHhoaaRUgGxl9cXLxz504IYIhuPYivBg0aBAYGlsJ9fOWAYRj4B4y1Y8eO6otO03R5Q2olto0kSZqm1YWvv79/VFRUXl5eiQFUGjie5zme5TiG53mWK35VtO1y7qB/3tU8mCM7kEUeyCYOZkkO59iczWl1P39Oruo2/IjjWfitaYJlWY7jMjIy6tevX5a92Lp1K8/zKpWqlMXiOJ7nFQqFi4uLjqdJpdKMjAzh+6YPILMzZ87oXqV+/foZhSb/t9TCdrx582bp0qUNGzZUFw3AOPpwH0VR6uLSwsIiKCjo+PHjAgkZbaY8z2dnZ0ul0g/bX7GxsSksLKwwX8Cv8vLyLCwsdLzF1tY2Pz/fLLgPKDwiIqIsqzd48OBSZZThlT84EBRFqVSqbdu2rVq16u7du/ARTdPgz+kZ0xd2V/ApWZa9fPny5cuXlyxZMmnSpDFjxlhZWcHXKi17wYGTRxKSLNX5u4VfZfKXEIFIAhEkIv73ThXLq7LRtUzm2hPlb/WkE+pbfEchK4RYhHhEcAhJTc3OAt8uIiLi0aNHUqlUR5qQoiiWZWUyWblkqL29fVZWlqajCTado6OjOdZfSCQSmqZFzVKaphmGsbGxMc5IWJYFLZWWlvbbb79t3Ljx7du3wCYEQXDvoX+4VT2iU1RUFBMTExMT07lz55kzZ/bu3duYgRmSJB0dHdPT0z9IjxA4xcHBwSB84eDgkJaWprlQwlvMIiHCMAxN04sXL96xY4duGQXcZ21tXeoz6crgQ4TQoUOHvvvuu4SEBJCY8FFl5LQFxgaue/LkybRp0zZu3Pjtt98OGTKEIAhYtUrYDopHKhJJUoo3/1f8BUcqJKQE8RyPOKRGYwRCNCIQQSAq/x6zLKPgZmvLnTLSiUccwdOmFiVlGEYikUyZMuWff/6RSCRKpbJUgVhe0QNkoE0RmmnVA4xcmyBmGKYsKX39xwDpQJVKtXbt2mXLlqWmpqpbn5UURec4DiI0PM+fPn369OnTvXv3XrhwYcuWLdUFQmXTrY71/wAUoaHoB7hPmyI0ApUaSgvu37//m2++AWrXvekMw5SF+EnD8iFFUcnJySEhIf369UtISKAoCtw1IywxbCTEXe/fvx8REdGrV6/bt2+DIDA4h/A8RxCSV4qdN4vG8pSCImieZ3hw9f7v/3jE8YjlESWnJOncqX8LBqhQFkIkQhxCvElRmEQi2bBhw5o1ayQSiUqlQhhmAtBGNE2fPXvW399/xowZqamp4KSWURDoz/uQkiRJMi4url27dl9++WVeXh5FUR98SSeG0QCm3n///Tds2DCISBlKsJMG5EOKojZv3tyqVat9+/YJKrBcTAg5eVINQpa+XLY5SZIURR07dszf33/lypXwHAMqYx5xBE/mMnf/K5pA0hzJkzzPlKbVWJ5TyQjpW+7y/cJvCILiCWQ6hTNgZ50+fXry5MlYeJkXgOCLi4u//PLLzp0737hxQ1CB5RITJbiPeI9yySlQhyqVavny5W3btj1z5gxN0wYUWBgfs7VHkmRGRkZQUBDkMg1IVKSh+DAvL2/YsGGjR4/OzMwEXV0WFSgk3iGBLyT/BMBfhK+VscQGvEOKooqKimbMmNGvX7+0tDTgTwOtG4FI/nHhIgWRSyIpj8qq7DmklFDUM9WfWaqLBKJ4njMRCqNpOikpafDgwSCzsNgyIy0IIZDAwMDly5eDCVhGFViCrUpwn1BHAPU1ZS+xAYtTIpEkJiZ26dJl0aJF8HyziLxhmCYE4hw8ePDTp08hzmfA59OG4sOIiAiIhZYl1iwkFUp8WS6XW1tbW1lZSSQSgiCUSmVhYWFeXh50u/j/2vt9gYxubhcipYcOHWrXrt2OHTvatm2rf8qQRxyByHdMQip7gKIpxJePvUmeUBGqJMVvDnQAIqpe3wA9FRQUBAcHp6engxGDGc+MtOD+/ftHjx6dnZ0NpQGlUyBJQk5IlPssLCwkEgnHcUqlMj8/v6CgoIRRCzm/UtMNKpUKeHzevHm3bt3avHmzra2tcVKGmjDfA45gfOADmpC4mThx4qlTp8pI58ZThEI8bfDgwenp6WUZnxClBA50dHT09fX18/Nr2rRpvXr13N3dq1WrJpfLJRIJPL+oqCg3Nzc1NfXp06e3b9++fv367du3MzIyhNpUwQLVZkfAIJOSkrp06bJt27agoCClUqlHvTWPeBYRZJrqkJIoliKaR+VVGxxNoEz2jIJ/IyPcEOKq8A4QsEVomh4xYkRCQkJlUBhGJW0c5EtWrlw5Y8YM4IVS946iKMHnQwi5u7s3bty4VatWvr6+derUqV69uo2NjUwmg6+pVKqioqKsrKyXL18+evTo1q1b169fT0xMVCgU/5MdpZXhQMZEIpHExMQkJyfv27evbt26VaILzde2g5Hn5OR8zEEalUolkUjWrl27bt26SpJRFVeEoGBiY2PDw8OVSmWpfCgwIXBgz549P/vsM39/f1dXV62Do2m5XG5vb1+7du22bdvCqZGMjIyrV68ePHgwLi7u5cuX6spVx1BJkiwsLAwODt6wYcPnn3+ul19IIMSjHFUCQRLv+6uVg0Z5xJMEUcylFzAPZFK3qm28BsJ0/vz5+/btw1rQjLQgmC9ff/31Tz/9VGoYBgIwgvXZoEGDvn379u7d28/PT2iWpgm5XG5jY+Pi4vLJJ5907doV/vjkyZNz584dOHDg9OnTBQUFJfhadKggxW7evNmxY8cjR440bty40gq5RSYO58S7detmpnsNU3Bycvpo27qyLCuRSE6cODFt2rTKi1dVkByBlHft2hUZGQlZBB3jgy2ELwQEBIwdO7Zv376Ojo6C2ShcKyGanC/RJt/JyalPnz59+vTJycmJi4uLioqC48zwFm0MKbQ+GT9+PMMwEydOrJhfyCOeQBSHVMUohSB4xCOEuPI/g2IRW8g9c4D/W6WBtZ07d/7www9Q0aDJhEis1w9GlYsGmqanT5++atWqUktRID4Jfthnn302ZsyYLl26CL17hPMPSK19jDbuo2nay8vLy8tr9OjRycnJ0dHRmzdvfvz4MahDHcFSlUpFUVRKSkrXrl2PHTvWrFkz4/iFoEWqV68eGxv7YWz9x9b4EGqvnj59Gh4eDoQq2rxJ/3xhRawMEKB///03aEE4qKvDqwNm69ixY1xcXHx8/PDhwx0dHcE+VU/FQ6EpoQHI/6sX1MBpGDs7uyFDhpw+ffrkyZPdu3eHt+jgLqH/06RJk6KioqRSaYVqZ3jEExxiWKRAPKpgK22CQARiCWWVC9Nr166NGTMGKEm06xLWgiYYJqJpeu7cuatWrZJIJDrqYiC3BDsbHh5+9erV/fv39+nTRy6Xw9FGgfsoiioL9wETwXmMunXrzp07NyEhYePGjd7e3kKXfN30lpaW1rt37wcPHhgzFc1xnEKhAKHBmi0+zsgHQigvLy8oKCgzM1ObwhMsOaMqQiDo8+fPh4eHA7Np04LARQzDeHl57dy588yZM7169RJOYQPvVWACQqmb8KguXbr8888/sbGxDRs2FO5m0hFToihq7NixBw4cADlSgcAohSQ0skEIvb9WovzaFCEJb1u1dlZqampISAh0Xi6xiWBweHh49OzZU/C2MaocUDKwbNmypUuXwmniUh3B9u3bnz9/fufOnX5+fhBBBY1VMe4TOrRB9t3KymrcuHE3btxYtGiRjY2N7pgnwzBAdX369Hnz5g0oaeOsG2X++Ai1IDQnGj58+H///ScasgLTLSwszNiKUHBUg4ODi4uLdYQihaLQL7744tq1a+DYgpbSp8uopkaE7CDHcQMHDrx69erMmTNLNPsWDfUQBBEREQHrW76Tjv/zAWlLwptDCKEKECiBeJbmZZZU3fea1dgUBjWBYWFhz58/15RHQpXazp07/f39sSI0HS1I0/S+fftmz54tKhdKhGEsLS1XrFhx/vz5Dh06AI+Ae2co7gNjFBrIffPNN1evXu3WrRvk47W9AszopKSkQYMGFRcX45ADhm6P65tvvtm/f79o+QIQ+eLFi6dOnao7FmhgRQhUW1RUFBoaCjWi2lQIfOTk5BQbG/vbb7/Z2dkJh9wrydwD19Pa2nr58uVxcXHu7u6QYtWmzgmCKCgoCAkJgXugyseNBI8I5CBphXjEkyzBE+WVICxBWBDuVlR90KzGt7Nomv7iiy8uXLigKU9BC6pUqqVLlwYEBECnSowqB1TH3L17d+TIkbrPDoHUaNq0aXx8/IwZM4QEYWVYM+rq0MfH5/jx44sWLYJgrLbXgTq/fPnyxIkT8VkdDFFA/H/79u2LFy+WSCSaRAJE3rdv37lz52ZlZen/xnLwBrDT5MmTb968qaPCED5q1qxZfHz8wIEDIYdhhCIxgSF79ep16dIlf39/WE0dYuXx48fjxo3TXekjwvuIRAi5SfpZIHsWsQRfXvlCqnjeje4nIezhCkPjexUrV678448/tNlZDMOMHj169uzZHMdpMyYwjGy+IIQKCgrCw8Pz8/N1NNWE7QsNDb1w4ULz5s0r1QAtoQ4hSf/NN98cPnzY3t5eh5EORLh58+bff/9dt2uL8XH6ghKJ5MqVK2PHjhXtDgFuT+PGjf/66y9DBRXIsg+Opuno6GhtAlSdD7t163bmzJkGDRqAKjJapRMwJMMwtWvXPnnyZL9+/SCnojvQVH5uJHjEWVC1PehRHMPz5TG0CUSwiLPg7Tzlk6oktiaRSI4cOTJjxgxRYxx8wU8//XT9+vVwIBqzpYmIBpIkZ82adffuXR20CsQ/bdq06OhoGxsb9St5jQA4p69Sqfr06XPmzJnatWvrKA2Fj6ZNmwaFM0ZLFmKYPqlTFPX69evQ0FDR4Dkk3RwcHGJiYuzs7HRUhBheEUIsMS0tbcqUKTpqVQV39dChQ3Z2djqCk5XtGnIcZ2FhERsbGxoaqsMvhEWfNWvWs2fPylWDSyCC5zkvi6/sUBMVUhGEpDS/jiAgCkrQHMP5SBdYUV48oSKMeJQenOAHDx5ERESIVoSCZ1y7du3o6GiJRIKTN6YjGmia/ueffzZs2FCqDfrNN9+sXLkSjOgqqbCAdu1NmzY9deqUp6enNl0I1FVYWDh27NjK6ImPYb6RD4VCERYW9uLFC9HyBcDOnTvr168v9HYwkiKEiP+XX3759u1bbQoDfNiuXbvu27dPJpPpn73U0ziFipidO3d+9tln2orZYN1zc3NnzJhRzjtcCETwMtKxhfVmOeeiQioCSQmthTMEQgSBJDxJKViVl2RiXfkknlcaMzsIq5GTkxMUFPTu3TvRW5AQQpaWljExMW5ubuCCYM40EdFQUFAwefJkHRXaoAVnzZq1aNEiKM6swgNnUIxdr169Y8eO1ahRQxstgYKPj4/fuHEjThZiCJ7JhAkT4uPjRSMf8MeVK1f26NHDsG0ZyDIO7vz583/99Zc2egVnwtfXd+/evaAFq1yMwgBIkty9e3erVq0YhpHJZLQGCIKQyWQHDx48efJkObmR5JHKjvbztzpixzViWCVHcASiCEQRiFT7H00QJEEiBaFiWKahZGZj6zU8wROIRjxlnNP0wk2qkZGRiYmJ2gpkOI6Liory8/MDSYrZ0hQABuXy5csfP36sLYQIiiciImLZsmVVrgUFgaVSqby9vQ8cOACXEosOCQTF/PnzdVjYGB8JQLH9/PPPf/75p7byBZVKNX78+ClTpuiI81WWIgQ7dO7cueomqqYzYW9vD0FbyM+bwsqCX2hpabl3714XFxeFQsGIAQ7bjh49Oi8vD35SNpeQIxDF84wd3TLA5qwXPZVmq6k4VsWzDOIYxLGIYxCnQoySZ1UM58D5tZMf8LFajhBLvPcpjVMpA6b37Nmzjxw5Inp0Erz5+fPnh4eHG5zCMPTRgiRJvnjxYsWKFdpKuiCt26pVq6ioKLBZTaT5CMRI/fz8oqKiQJ1rDgxyLhkZGYsXL/4gr9XFKJcWPHTo0Jw5c7SVLzAM06lTp7Vr11ZG8psuizsYGxt7+fJlbQ4TDDEqKsrb29toXQTLrgsh77Vnz55du3ZpO/IBN128fPnSx8enzH0KKIR4gqB5xEpJp8bWq+qyX6Qq4rLYMwXoKcsXEYhHiJYQztXI5q7S7i6yLiRhxSOegDUneOP02oZOj5s3b16+fLnoEWwwvkJCQhYuXGhq24cVIdjIeXl5ojYyKI9q1art2rVLLpfDOV3TGT9YXaGhoVevXv3ll19EpwDKftOmTZMnT/b09DSFYBJGldD5vXv3hg4dKtrlCsS4p6dndHS0kPYyqiIEd/Cnn37S9mLQgmPHjh00aBDIXFNbZWgKHBgYGBgYWEbdWeZnw9W6FBxSt6Lqe1nWR/w0xCtZohghnkA0iSwQQUIEFO5vUv+tEewsiURy4cKFCRMm6LCzWrRo8eeff4IM+tiaGZqydKAo6vnz51u2bNHWUx4ExK+//lqvXj3TNGKAwBYvXnz69Olbt25pxj+hqKeoqGjFihXr1q3D0dGPkM4JgsjOzobyBdECGYSQlZXVvn37nJ2dK6lLLanbHSRJ8uTJk9evXxdlRSDrmjVr/vzzz6DVTXOtQZ0zpUG/N1AIcTxieMQikqYIW4qoRhJWPIF4pOQIhic4wrjXLcGOpKSkhIWFKZVK0TJRjuNcXFz27dtnZWWFcAcZ0xMQ69evz8/PB2NOU8ewLNunT59hw4aZrCsPNX4ymWzdunUgvzQtLXBkd+zY8fr1a3yU4qOCcG9JRETEw4cPdZQv/Pnnn3AutpLKF0jdRIwQWrt2LdKS64bIzJIlS+zs7AzS+bTyILT21gH9X0Ig4dAkjxCHEE8ggkASEpHG7yCDECoqKgoODk5NTRW1s+Co9e7du+GWOKwFTUpA0DSdm5u7detWURsUWA+aqOlo42I6TqG/v/+oUaNEi8nBKYTJIu1dGzE+PIABN2PGjKNHj4qWL0A4feHChSEhIZVavkDqMEhJknzy5MmJEydE67bBIG3duvWQIUOMfHTXtEEiRCJEvP8vofZv44lRUGxjxoy5du2atkJkhmHWrFnTqVMnXCZqaoD9OnDgwJs3b0QLuMCyGTduXIMGDUzfiIEpfPfdd9WqVRNNZIJ42bp1K9xsigngI9GCEolk06ZNcI+KaPmCSqUaMmTI/PnzKzvvpksRIoT27t1bXFwsGpkBzJs3D2eVTFCM0jS9aNGinTt36ihEnjJlCtzOiI0YE9QcCKG//vpLWySGZVlbW9tZs2aZuDsoTIfjuBo1anz++eeiJ/3B7H748GF8fLzuS7YxPiRf8Ny5c9ByVlspe+vWrX///Xcj5N1IHQENnufhQkvRuxBZlm3RokWvXr2q9uw8hiiFxcTEzJs3T9QXBArr3r07tCDBe2dqAK3w/PnzCxcuiF7+DrwZGRlZo0YNcymzhFjupEmTrK2tRZ1CmMXevXsRvgv6I6BwmqafPXsWGhoKXZBK7DiEG93c3Pbs2WNpaYkq/0ZiUttACYJ48ODBrVu3tKUoEEJQi4hj+qbmC966dWvEiBFATKKFyN7e3rt27YI0IXboTVBMIIT++eef4uJiaCUvusvjx4+vjDryynMKWZatVavWoEGDtDmFMGuFQiE6a4wPA0J3veDg4Ldv32orE5VKpbt374aOtUYw9UgdrHjq1ClRjwFUo4ODw4ABA0B74901HU8iPT09KCgoPz9f07IGCrO1tY2NjXVwcDCLqNpHCNimY8eOifpG4A62b9/e19fXvHYQnMIRI0YgsYoYML6fPXt269YthEtmPlwtCIpt1KhRN27c0BayYll23bp1gYGBRgtZkTpY8cyZM6Kfwsh69uzp5ORkamd4P2YKgzDa4MGDk5KSdBQib9u2rXHjxqbTAAijxD5SFFVQUHD16lWkPUgYFhZmdtoC5EZAQEC9evVEI7qg48+dO4dwdPQDBSRuFi5cGB0dreMauOnTp48ePdqY5QukNlZUKBQ3btwQZTag0f79+2NiNSkKg9siT58+rYPClixZ0r9/f1wgY8qKECGUmJiYmpqqWa0NwRiZTNa9e3dkhuc+WZaVSqW9evUSHTzM/eLFi6jyc0IYVSKjJBJJdHT0ggULRK/bhfKFvn37rlixwsjlC6Q2Vnz69OnLly81FSGworW1dfv27Q11FxSGnoDa4jVr1qxfv15UC0J18vDhw+fMmYO1oCkD2O3mzZuiiTRgt6ZNm9atW9ccI9ug3kCLa5rR8Jc7d+4oFAodleoY5ghIbN+4cWPUqFGQMBYtkPHx8YHrdo1cvkBqY8X79++LNrOBwTVq1Mjd3d2McvUfNoVJJJLjx49PmzZNx3W77dq127hxYyU1KMIwLP777z8diiQgIMBMzxjA+Fu2bClaOwqS8eXLly9evEA4OvphmXcURaWlpQUHBxcWFmqGOuCAjZ2d3b59++zs7Ixv5Gn1DO7fvy8aoIARt2zZUlDyeJurnMKePHkSHh4OgkO0TLRmzZp79uyBG7Kw7WLKAP5/+PChDp+pbdu25js7nuerV6/+ySefQONG9TmC+FOpVElJSV5eXoZShMI1ZKZguMMAhAtmPwaShtoFlmXDwsKePXumGbKCyCLDMNu3b2/YsGGVhKy0vi8pKUnHz3x9fbHM0txs3QxgcBsHGDsvL2/QoEFZWVma7qDQ6XHv3r1wRarpu4Msy0LYxCzEBCy4oYpWQBMwDKPNJYKKOx8fH2S2jWHBdPPx8bl+/bpmD274S3JysgE9QoqiZDKZaS4FTJ8kyQ81x8TzPKQGP//883PnzunoILNixYo+ffpU1c0NtDabFBKEmrQIO+ft7Y1wQvv/qhwj6xiwc2maHjZs2J07d0RTg5B83rx5c9u2bc0iNUgQhJ2dnRkFb2GocPesofDu3bvMzEwkdvqF53l7e/saNWqYL/fBpOrXr6/jOykpKQZ8V1pa2siRI8t+1ahh6ZkkSalUamlpaWNjY29v7+rq6ubm5u7u7ubmZm1tra7/oJD7A9OIoAVXr169adMmbdfAqVSq0aNHz5gxA75cJeOktTnv2lgR6p7d3d2xIlR3y7Kysi5evKhtQSAc5OfnZ8BFAy349ddf//3339rsLIZh5s6dO2zYMNO8IUsTSqXy8OHDNjY2ZuQRUhR19+5dA9JSdnZ2Xl6eKGPyPO/s7GxnZ2fuXOPh4aHD53v79q0BFWFubu6WLVtMavo0Tbu5uXl6ejZp0qR169atW7du0KAB2KkQW/owEvlQvnD06NHp06fruAauQ4cO69evr9p4FS3KigzDiLIiwMrKClgRK0Ih1HP37t1+/frp+Frfvn0PHTpkqM0G927btm0//fSTVCrVpgUHDBjw008/mcWl8yCzcnJyBg4caKYhAUM5HPn5+bChov0QHB0dK+luUqMtFELIxcUFaU+CZmdnG1a8mA79w3lfhmFevnz58uXL8+fPI4QkEkmTJk169erVv3//li1bgogw9+pukHWPHj2KiIgQJq7+BShf8PDw2LNnj0QiqdryBfGFVqlURUVF2n5jYWEhl8uxCizBaRRFiUpD0Elw4Z8BfcErV66MGzcOKkI1C5EZhvH19d26dSvoaTMSmuab+jKUNVBYWKhDs0IY1twLtoVZiH4KK2BA6HfbaGVZA3BCgOd5lUp148aNGzduLF68uEOHDmPHjg0JCZFKpUIG0ez2F7Ld7969GzRoUHZ2tmj5AqiSmJgYNze3Ki9fILWJWh02qVQqNYs4m5E3ntUJQ9VTgGJ79epVSEiIQqFAYmWiHMc5OjrGxMTY2tqa3WkzzjxhcKmtTc9B3Yf5Hi2AeemuXlEqlR92wAncI5ZlGYaBYyRwYSo01omMjGzTps2ePXsgZWhqWryMwpAgiGHDht27d09Hl6uoqKiWLVuaQut/smLzxJqvCvmnuLg4NDT05cuX2spECYLYuXOnt7c3vlwCA8Ms+JrjOLiHgaIoiqJu3boVFhbWv39/aJdoXroQQlZz5sw5ePCg6HW7ELKaN2/ekCFDTCRxI64IKYqSSqWaRhmoQJVKZXZGyocBCCBMmDDh0qVL2q7bZVl21apV3bt3xx1kzBQQbtFmbpq7twTzglloAwifj5bH4ZAMRVEHDx5s06bNgQMHROstTRMgdrZs2fLzzz/r6PUYEhLy/fffm46MorURooWFhbbfFBYWFhUVVatWDcss41PY0qVLt2zZouM4zhdffDF58mSsBc0RoN7gAjZtyM3NReYfNoRaPG15UGtra2S4yJPxjzZpswAElPplCLZTFJWRkTFgwIDVq1dPmTLF9Gu/QexcvHhx/PjxotfAgS/YokWLLVu2wAEEE6FkWpQVKYrScTSqsLAwOzvbzc0Nt1gzMoUdOHBg7ty5Oo4MdurU6ddffzXrPmofc7GMoAakUqlSqdRsvIIQyszMNJfLeHV4hHBAQlN6wF+gKN2AnWVMKoIFmT8Ih+qeI7iGBEFMnTqV47hp06aZsoEL1+2+ePEiNDRUoVBoHtyE8gUXF5e9e/daWlqaFBnTovMhSdLZ2RmJhUZhMqmpqT4+PjhZaLRoCU3Td+/eHTZsGKy/aCGyp6dndHQ0ZKHN10D5mC+ig12zt7e3sbGBg7yaKiQ9PT0nJweukzTfXYYj89rG7+rqaqj1hJOX06dPryqtr1Kp8vPzc3JyXr9+/eLFixcvXuTk5AhEDrkMHYIUeJmiqOnTp9euXXvgwIGmaebCFIqLi0NCQl6/fi1avgBqb/fu3Z6enqam0bUqwlq1aiHtvUafPHnSuXNnrAiNoxhIkszKygoODs7NzdVWiGxlZRUTE+Ps7GzW7iBN097e3mY0fjANCwoKdLckLBdsbW0dHR0zMzNFI4dZWVmvX782X0UIY3706JGO79SuXduwivCrr74yEWpJS0u7f//+xYsXjx8/funSJXBVdZfDQDQVLrNt1qxZnTp1TDAkAMb6uHHjrl69qi01qFKpNmzY0KlTJxP0a7WOpl69ejp+pq07PobBOQeIPiIi4uHDhzqColu2bGnWrJn5pgZBZjk4OMTHxzs4OJhL9A/Mjvj4+A4dOmh2zqzAIkB8ycPD49GjR5rxGDCD7t+/37hxYzM1Q2FbExMTkZZ76hFCderUQYbLgzIMo1KpgMCq0HQAx87Nzc3Nza1Tp07ffvvt7du3t2zZsnXr1qysLAgkattTODSVk5MzadKkI0eOmNrWg9hZvHjxX3/9pUMLTpky5fPPPzdNGUVqM9k++eQTpL3XKNzZi0vzjUNhM2bMOHbsmGghMvzx+++/Dw4ONosOMqUqfqE9vzkO3iABAIH7tKXQrly5gszzFBPsb2pqquj1GmAHSKVST09PA9IAQRC0CQBuWBSOD/I836RJk19++SUhIWHMmDGQ79Bh/IHJFRcXd/ToUZM6XAgyav/+/d988422UnaGYXr06LFy5UqTPdClVRE2bNgQZiV6guLu3buvX782YFspDFEKk0gkGzZsWL16tY4y0SFDhsybN89cuomWRVaaFypjEZo0aaJD1168eFH02l7Th2BG5+XlaV69C6KmVq1aNWvWRB/igXo440tRFE3ToPUZhvHw8Pj999///vtvJyensgRCFi9ebDpdMiAievv27eHDh4uWL0C8ytvbe+fOnZAmNM1t1aoIa9euLRqgAA4sKCi4cOGCcNEXRiXZWWfOnJk0aRIQU4kvwB9bt24dFRUF8bQPSViYESojctiiRQvRq3dBkfz333/Pnz/XvN3UXJzmEydO6PB3mzRpIpVKNU3wDw/QTQbUYf/+/c+ePevh4aFDF0IR6cWLF+EGqyqXvTDUjIyM4ODgvLw8TYIET6latWoxMTEmnu8QV4Qsy0qlUrh9V3PoQKAHDhz4eO6WND6F0TSdnJw8ePBgMLJEy0Td3Nz27t0LJz7xRnwwdgBCyMfHp2bNmpo5LTBDi4uLjx8/jsytwhYGr1Qqjx49Kjp4mGxAQAD6mNpXgTpUqVSNGjWKi4uzt7dH2g8RQSpx165dVb5EQkRkyJAhjx8/1tFHbevWrb6+vibe5YrUYbh17txZm2GCEPrnn38yMzM14xsYBrGaCwsLg4KC3r59q1mCIVxyFh0d7eHhAXYiXrcPRhGyLGtpaenv76/jJPju3buRuZ25BDK+dOnS48ePRQuLwAsMDAxEZnuctMKA3EejRo3Wr1+v4/gTLNqJEyeqvDgcBjB16tQTJ07oKOJbunRp//79Tb98gdQRn+ncubNMJhNNE1IUlZWVFRsbK+hFDANSGEmSI0aMSEhIEE0+A4WtX7/+008/xd1EP1RLqGfPnqI5SODHCxcu3Lt3z+yiowRB/Pnnn6J6DnwdLy+vpk2bfoSKEL0vfAsLC+vRo4c2PQfb/ejRo2fPnlXh7oNi++2339auXSuqBWEuw4cPnz17dhVet2sARcjzfL169Vq1aqUjMQu3KWJ3xICA1OCCBQv27t2ro1PfjBkzRo0ahfuofZAAhurWrZuVlZVoqgwsoY0bN5pRtRrkh168eBEbGyua/oRZ9+zZU7Q6+uOJB/A8P3v2bKQ98gnh5Tt37qAqio6CYjt58uTUqVO1XberUqnatWu3ceNGcznWrKtaFyEUHByMxPJPoP8SEhKglhc7hQbUgrt37164cKEOLdi3b9/ly5djX/ADVoQcx9WsWbNjx45CPw5Np3Dbtm2pqan6H140mptLEMS6devy8/NF8ykwi5CQEPQRJ7yBozt06FC/fn1tpSWwOKLnT4xj0NA0/fTp0/DwcB3lCzVr1ty7d69MJjOXOhJSt1kaFBRkbW3NMIy2Eq8ffvgBFgJnCvUEFCJfv3591KhRotIN/AAfH5+//voLvb/VE6/bBwnY/aFDh4qylXDr6YoVK8wiOgoyPTU1dcOGDdrcQY7jGjdu7O/vb6YnQwwoByQSyaeffop0xoehR53xrRmEUH5+fnBwcEZGhmj5AkEQcrl879697u7uZhQv1KUIQbH36dMHiZ2dB5/333//3bFjh6iDjFEuSUFR1Js3b4KDg4uKipBGTR3QnJ2d3d69e+3s7My67TJGGT2Dvn371qxZU3SvgWDWr1//6NEjqM0zcfImCOL777/PycnR7MUsSPyRI0ea3d17lQRfX1/dX4BWtMY0heGwHEmSw4cPv3XrlrbyBZZlf//997Zt25pXyKp0YTpp0iSkpVAbwh1fffWVNvo2KVZkS0MVRo04jlOpVGFhYc+fP4ejRSXsLJAU27dvb9SoEQ6KfvAgCIJhGCsrq1GjRokm6YXS4lmzZon6WKYW6rh69WpUVJRoqAPGb29vHxkZiRDCaW+EUPXq1ZGWMBv8saCgwMhDgsTNvHnzYmNjRfO4YMR89dVXkZGRZtflitRtlnIc1759+08//VT0yDbYqi9fvvzyyy9NquuPaCiJKg1VNTaQFF988cX58+d1UNiyZcv69OnzAfRRwygTZ5Ikz/Off/65ra2taMkMkM2hQ4e2b99usgUmwjW8EyZMgBFqCndIGY4YMcLFxeVjOEdfqg2EEJLL5aWqJWOOChpX7dixY9GiRTrKFwYMGLB48WJzLOKjy0LH3377bffu3UXNE+DGqKionj17BgUFmeASQAj36NGjmzdv1hZEIghCqVQuWrSoUaNGRo46QgnWypUro6KiRO+hhj+OHj165syZZlGIjGEoRcgwjLu7+7hx45YvXy6afYBQ1eTJkwMCAurWrWuCFXogH+bOnQtngTQFKOQ4ra2tp0+fju83La+FYbRNlEgk//7775gxY3Rct9ukSZOtW7dC0N7s9rEUpQWao1u3bl27dj158qQoN4LmGDt2rK+vb/369U2KG2EwT58+jYyMzMrK0vFNHx8fT09PI9/kB4otLi5u5syZ2gqRGYb59NNP4aQKjoh+hE7hrFmzNm/enJOTo3lYAjRHTk7OkCFDzp07B66V6cggMItjYmJAkYs6MVCLMHHixFq1amEKVxdcZVGERthrUGypqanBwcHFxcXartt1dHTct28fRC/MsXyhrCNesmQJTdOibAbKIzs7Oygo6N27d6aTugcNnZ+fHxISkpWVJZVKod1tCchkMoqioqKiLCwsjClHwFh++PAhpEa0FSLXqlUrOjpaIpHghnYfoSJkWdbV1fWrr74CeSRK5DRNX7lyZdy4caLWetVqwYSEBG1V0IIMrV69+uzZs3H9l7qGy87O1q3nZDKZEfxCEEpKpTI0NPTFixeasl3ouLtr1y5vb2/zLV8gy8iNfn5+EyZM0MaNYMrdvXs3KChIoVCYQkk3qGeGYUJCQiAso1Qq4QIUdSCEFArFkCFD2rVrZ0yDFNKWOTk5gwYNys7OFi1ERghZWFjExMS4ubnhxgUfJ0D0TJkyBbo1itIAqJytW7d+9dVXpd54bkwtmJSU1L9//9zcXG0iG9yLJUuWmO89w5UEHRcXwypBS1IjGOsURU2YMCE+Pl7bFUssy65evbpbt25m3d+jTLIVxPT3339fu3ZtbRIZ/JtTp04FBwcrlcqqPWUPg2RZNjw8/NixY9pqsmFeLi4uP//8szGDosKtHUOHDk1MTNRGYRzHRUVFtWrVCpeJfrSAcKhUKv3tt9903GIDMmjJkiXff/99letCqOdKTk7u0aOHqBsh6Hi4pm7YsGE4KKoulFAZrpx0d3evbI8QiGrZsmWbN2/Wcd3uxIkTJ02aZO5FfGTZudHOzm7Dhg06Oq7Bwh0+fPizzz7LycnRlhUwgjUKF0UNHDgwJiZGR0EdKMI1a9a4ubkZMzIDRsPs2bMPHz6srUxUpVLNnz9/yJAhuEwUO4UMw3To0AGqpbQpDOC+77777ssvv4Tr7oxvifI8D2nvu3fvdunS5cmTJ9pOGIMvKIgU7Auqx7GSk5OvXr2qO67m5eVlBJ/+8OHDs2fPFt1EUI2dO3devXo1CDTztj/KxY09e/acNWuWDhcYPjpx4kTnzp0fPXoEi2U04xT4kKbpZ8+edenS5fDhw6J1mADQQKNGjQoNDTXmRoJi+/PPP5cvX66tTJRhmODg4IULF+Juohjo/TnlH3/8EcID2kgCyHj58uWRkZEFBQVGtkQ5joOE5dGjRzt27JicnKyjzwbYoOvWratTpw7ODpZQhJs2bSoqKtJ2sQ8saePGjVGlFcsAISUmJg4dOhRMFs3yBYZh6tWrFx0dDVEKczdlyPJy45IlSzp16lSqLkxISAgICIiNjTWacQp94IAP27Vrd/XqVR1dKqAtbLNmzX799Vdjpt/gOE58fPz48eN1lIk2b958y5YtZlqIjGFwgKCRSqW7d+92dHTURrGCQ7Zjx47AwMA7d+5AgL2yuY/neZVKBUd1ly5d2qdPH7igTdt7wQadMmVKeHg4DvuXUD/Pnz9ft26dttQSBOfc3NwaNWqEKueODmhxB8WPUKus2eUKIWRtbb1v3z4nJ6cPw44hy8uNJEnu3Lmzbt26uqM0FEVlZGQEBQVNmjQpOzsbrJtKYkiWZcEUzcvLmz17du/evVNTU3WYw0BkTk5O0dHRVlZWxjRnZDJZSkpKWFiYUqkUtbMgZ7lv3z4rKyuEr9vF+L9E6+npuX37dqFUT0fI4caNG+3atVuzZg1cagg8UkkxGIIgJBLJw4cPe/XqNXfuXBibNmaHKEj37t1XrFiBtaC6LwjrOXr06NzcXG33ioBl3KFDB2tr68poPiAIpcjIyAcPHohetwtiasuWLc2aNftgdpAsLzdyHOfm5vb333/b2dnp8KXgI4qifvvtt9atW+/atQsYEtShQYKl8Chwm0iS3L9/f5s2bZYtW0aSpI5SHSAdqVS6b98+OPVoHHMGKKyoqCgkJOT169eaFQQwbIqidu/e7enpqa1EEOOjhZCeWLduHVSXaJODIJ7y8/OnTJnSpUuXS5cuAY9AK0FDcZ8QgyksLFy6dGnr1q2hME30GkVBCzIM07hx4+joaBgSNvXQ++tEKIoaN27cqVOndDjTsLZwR0dlpJxA7MycOTMuLk5Hl6sffvjBNNun6CugywWVSsXz/NmzZy0sLEp1zwV7oWPHjnFxceoPYRhGuMij7OA4Dk4+CH85efJkjx49hH3SrciB9/bt2ydMRE/ASOLj47U5cDCk4OBgnufDwsK0DRJaxqxfv95QA9MTMIZp06ZpGzDs7K5du8oyYNhopVIJSX5NmoGlc3Z2zsrKEr5v+oDdv3Dhgu7dHzx4sPBlg+zLjz/+CA/XoUiEC+4Jghg6dOitW7f05z6WZeG38H8LCwujoqJ8fHxKMLs2LYgQ8vT0fPbsmWAQ6wMY/7t375ydnUXXH8isQYMG8E0TJCqGYWBD8/PzhwwZoluCgezy8PDIy8uDjGy5FiovL8/NzU3HQlWvXp3n+T/++EMQR6I7GBERAbxsItx35MgRbbQHA4aGvbplFNKHG48dO2ZpaVkqA4D6gX+3b99+y5YtIOyETVKpVMBd4OGVABz+g++oM09OTs6OHTs6deqk+RYdZEQQxI4dOwyobMqoCMeMGfPbb79pI3T44+TJk01EC2JFaLKKUFjtBQsWCLGysliiEokkJCTk2LFjCoWihCAG1iuV+9THkJycvGzZsk8++USYpu5hwDrUrVv30aNHhloK81WEsKqCNDt79mzTpk3LaEn88ssv5ZUSpSpC+EvDhg2PHj0KjUdEr4NGCLVp06aoqAhIBStCXrAITp8+Xa1aNVSGnvEQCYF/u7u7jxkz5sCBA2lpaeV9b3p6+uHDh8ePH1+rVi2B3EuNU8PwpFLpnj17DKtsSlWE8Ed3d3epVCqqqmFs3bt3B/YwEXbFitBkFSHYjnAUHaldTlIWdQjybs6cOefOncvLyyvvq5OTk7ds2TJgwAAbGxtNvtYGcC98fHySkpIMyH1lV4SgdQR9b3wIxkQJGrhx4wY0lipLNIsgCC8vr/z8/PLqoVIVIcDBwcHBwUG3y5iSkmIQb97UFGHFI7wQQe7UqdOZM2eCgoKSk5N13yUGUW/YztevX0dFRUVFRTk4ODRp0qRVq1a+vr5eXl6urq729vYWFhYwK5Zli4qKcnJy0tLSkpKSbt++ff369f/++y8jI0OdvUutwYGBOTk57dq1q2vXrkYObUMo//Xr1zoSP97e3pBGxQUyGKUCMnMqlWrOnDmurq5jx44Fki6V+yBJn5iYmJiYuHTpUg8Pj2bNmrVs2bJRo0Z16tRxdna2tbWVy+XwNZVKVVhYmJWV9fr164cPHyYkJFy/fj0xMRHuyxSeppv7IDyrUqkCAwP37NkD90sYObEkhIhNB48fPz5z5sy+fftOnToFtQKl3t4Dy7hmzRorK6tKqmyAbsyadTpC0XJ0dPSH2hJWL4oE3mvevPmFCxeGDRt2+vRp4A0d9WnwkVA4k5WVdfbs2bNnz8KnMpnM2trayspKKpUihBQKRWFhYX5+vkKh0Iy1lqUuHF7EMEzTpk137drl4+NTVQleUG+aFMbzvI2NTUxMjIODA+6vgVFeS3TEiBG1a9ceOnToq1ev4C86aigEYxQKZ1JSUlJSUg4ePAifWlpa2tjYWFhYSCQSjuMUCkVBQUFeXl4JAQ0kWhbug6IPhmFGjx69du1auVxeJRSuVCqfPn0K5+GMbGWCoQDW/Js3b5KTk+/du/fff/89fPhQkGllqemVSCRg9/Ts2bPy1lBURgnG+u+//96hQ4cP9WSzvlMCXVijRo1//vln9uzZK1euFP6omz7gC0LXKFCfCoVCoVDA5csldkj9a4Ayjo1hmJEjR65evdrGxqYKOyBokpegpLdt2+br6wtHDLF8xygv93Xq1OnSpUtjxow5ceKEcFhCx68E9hGy5iCvCwsLCwsLRYNy4H9AhK0sh6AEn9XCwmLZsmUTJ04EzjWyFoRpJicnw6m7KuF6bZd+C9GsUtcTtGBISMiSJUsqtcpd1IQCGps1a9bIkSM/4P4etEG4EUj8l19+6dix49SpU589ewaqqyyXiah/RzM2CHtT3jOIQCsMwzg7Oy9btmz48OHo/X0iprP0oAV/+umnAQMGYC2IUWHuY1nWw8Pj2LFjS5Ys+eGHH4qLi0sNzKjrCW3cJ4jFMtqdJRxBlUrl7++/Zs0aPz8/EN9VdRyI5/kSISXjQ739SrnsCZIkVSpVcHDwjh074Oi6MZ1akFH9+vVbtmzZh33okzTUNoN1069fv2vXro0fPx6971xerrUTUrsCdJxJ0jYSOKLHcVx4ePi1a9eGDx8OqV2TOpYHUaxhw4bNnTsXX7eLoae0AkX19ddfX7p0qXPnzkKLiXLRfAnu+/8FdeUZCZi/tra2ixcvPnfunJ+fHwjQqs18E1UK9D6SDAEqWNuymDigL6dOnbpnzx6oyzWyFmRZtmHDhtu2bTO+DjZLRYjeB/qgY8v69esvXLjQtWtXcPxJkiy1utogbweJwLJs27Ztjx49unPnztq1a5sCH2pSmEqlateu3caNG3FeEMMglihUWzRv3vzUqVNbt26tX78+yFyapiubwCAQCt1keJ6PjIy8du0a3AkFA6jy9ana4saKLSbDMNWrV9+5c+eqVavUXXajURTHcfb29vv27atWrZqpORKmqwgFEQ9WTEBAwIkTJ/7+++9PP/0UjsDzPA8mqmG3E7SsEIv38/Pbvn17fHx8z549wfIytaA2mMw1atTYs2ePXC7H1+1iGAqgeHieHzZs2PXr11etWuXl5QVnBkC8GlaWCdYnpPxJkhw0aNClS5f++usv6NmEKqcZ5gdszYAoA9t97Nix165dCw8Ph5U0ppQQajJ27NgBBYYfvLFOVsYiCsHJ/v37nzt37vjx40FBQRYWFkJYAKzUCjMJxD/BbgItK5FIevfuvX///mvXrkVERICyMUFfHtSeTCbbu3dvjRo18HW7GAYXpuBM2NjYTJ06NSEh4Y8//vD39wfxCgEu4L6KsQaISHiCYH06OjqOGzfuypUrMTExbdu2FRofYguvLItZQpRZWlpGRkZeuXJl06ZNNWrUqJKAFhTIrFixolevXh/JNXB05TEket9xtFu3bt26dXvy5ElMTExsbOzNmzfVa0qFbRbCCOrBBCGBL3ynRKlx8+bNBwwYEBQUJBSGVSA3aUAlV6rHzDAMyCazKMEqkeoo73zL/kBzP0OpbTV0N8iuPEEG2T5ra+tRo0aNGjXq0qVLe/bsiYuLe/z4scA+6sXYSK0wTTf3CWU4VlZWAQEBQUFB/fr1g5Pa8PeqYj1TJh518i6xmMJ3WrRoMXDgwMGDB0PTCSGpZOSFgnLfsWPHTps2zVzKF3TLqLIQRuUKYvUD715eXnPmzJkzZ869e/fOnDlz+vTpGzdupKSk6K6eEo2w16hRo3nz5h07duzcuXOzZs3Ua7GqRAUitSSE7q8xDDNnzpzhw4ebS5koxLRF73QEa6a8KRDo76C56fAcbZdHmjh07D7MyPg3VAtHdcELbNeuXbt27ZYuXXrt2rWTJ0+eP3/+7t276enp5eU+iURSr169Vq1adenSJTAwsE6dOoL1WZYGN5UEoWmIKVOI6B7VrFmzadOmgYGBnTp1atGiBYgyoQd3lSwUdD8QGrubC/eJsljZuc8YHolwAhd4slGjRo0aNZo0aVJBQUFSUlJiYuL9+/efPn366tWrt2/f5ufnFxcXw9BpmpbL5VZWVs7Ozu7u7l5eXj4+Pj4+Pl5eXtbW1uoSuQqLswUBYW9vD4d2tbnILMt269ZtyZIlZnQcx8bGxs7OTiqVahITlPzIZLJyPdDJyendu3eal29AaMjJyckcnUIduw/2NbQhrBJLWZ37ZDJZ+/bt27dvjxDKzs5+9OjRvXv3Hj58+PTp09evX2dlZRUUFCgUChDEEonE0tLS2trazc2tZs2a3t7ewH1169YVbDh4bFVZn8IcnZyc0Ptb702NNsCrk0qllpaW1apVc3FxqV69uqenZ/369b29vevWrQsXrqmLsso7L+/k5ASOpjbFzHFcrVq1oqOjId9sFswolUrt7e01b4wC7lMqlba2tqVGmwjjkw4wjzaTp7i4GFgRRK1MJpPL5doeYjqXIyuVynfv3ulWxhzHOTo6mlEMkOf5goKCoqIibfPiOM7GxkZ0g7R9/927d9ruUYPKtGrVqpldZl6lUsGlm9pmLZfLheacVbuhEDgRNRw5jisqKlIqlcCeYIZCj6cSEM4jmUKGm2XZnJwc43eNKZctIpVKoTJOhyir7MWEiwp032IIR1/kcrnJrqcmFApFbm6uttXjeV4mk5XKfUQV2lDCuSXBdNIhcNW/ZtaVlmZEYRgfMASliNSShbq5T0fCGKMCa26yi/kRyijC1IIJoq3IzIXKy2Iemh3rGnZGBn/gh7pQJjIFcwldmIGoNYGVLONCfYRiijALGsLAwMDAwKgk4ENsGBgYGBhYEWJgYGBgYGBFiIGBgYGBgRUhBgYGBgYGVoQYGBgYGBhYEWJgYGBgYGBFiIGBgYGBgRUhBgYGBgYGVoQYGBgYGBhYEWJgYGBgYGBFiIGBgYGBgRUhBgYGBgYGVoQYGBgYGBhYEWJgYGBgYGBFiIGBgYGBgRUhBgYGBgYGVoQYGBgYGBhYEWJgYGBgYJgL6LJ8iWVZkiQJgkAIcRzHcRz8nSAI4e+GBc/zHMfxPP8/dU2SJGlInQ1TKPFMeKmhZgRPE7c+DL1oLMsKa0UQBEVRwt+Ff+v/Ch3DFl1PA26Nwd9iNEpTH7w6PVAUVRmMUxmcyLKs5rANu+PC+qvDUKRb2VusgzV0c43+hGRACay+y5UhskR3uTLUR4ntVpeHOkCIkqC2mRAEoTluA0pbHQ/keZ7n+UoSUjA7Az6c53njiDmjbYqOeRl26YwJ0VUyuPACzhR9UWXskRHWx8SHrQ6GYWiaNoJ++vCYvQppzPi6o3RFqFAozp8/37RpUxcXF4TQ5cuXT5w48fbtWwcHh8DAwC5duhhWFMJMWJY9fvz4xYsXc3JyXF1du3bt6u/vb5AXgSi/du2aTCZr0qQJ/F94aWZm5unTp/39/WvWrKmPJoN1f/bs2aNHjxiGUV9hgiBomvbz83N2dtZfWQqrcenSpVOnTqWlpdnb23fs2BE2JTs7OyEhoXXr1tbW1hV+F/wwLy/v8uXLTZo0cXV1hVmU+MK///7LsizskT5bgxB68OBBdnY2PEp4EbwlMTHx3bt3/v7+BrEz4CHp6elxcXEJCQksy37yySe9evXy9PQ0oCkjPCclJSUuLu7Bgwc8zzdq1Oizzz6rXr260QymCg87NTX1yJEjd+7c4Xne19e3b9++1atX5zju1q1bEomkcePGJeihApySkJCQlpZWwmOgabpTp05yuVz/WWRlZR07duz69esqlcrT07NPnz7169c31BZzHHfu3DkvL6+aNWtqssaZM2caNGhQo0YNAzL7jRs3Tpw48fLlS0tLy7Zt2/bu3Vsul+upToqLi8+ePatSqQiC4HmeoiiSJFUqFcyI47i2bdu6uLhUbBawrffu3Xv9+rWwxRzH2djYNG3a1MbGxlBcICzCyZMnz58/n5WV5eLi0qlTpw4dOpSuO3jtANJMT09HCF28eDE1NdXf39/CwqJp06adO3du0qSJRCJp0KDB+fPneZ4Hia8n4CFHjhzx8PCwtLRs0aJF586dGzVqRJJkhw4dnj9/DoEa/V8REBAQEhLC87xKpVIqlTzPP3r0yNXVtUePHunp6SzLwtwrBpVKxfP8woULEUKOjo4272Fra1utWjWE0JEjR/RfMfj5nTt3/Pz8SJL08fHp3Llzq1atbG1tfX19Hzx48OzZM4QQiLAKLxr88NatWwihDh06sCzLMIywOPDpvXv3EEKNGzcWaEafdZs4ceInn3xSYszw0bhx4zQ/qgA4joMnfP/99xYWFs7OzgEBAYGBgR4eHhKJZNSoUQUFBXrORVgfjuOUSuXkyZMlEkmNGjUCAwMDAwNdXFxkMtmPP/4I+6j/iwwLWB+O47799luZTObk5NS+ffuOHTvWrl3b1tZ2wYIFPM936NBh0KBBwtboQ12BgYHqnGJra2tjY1OzZs03b97oswvAICtWrLC2tnZ0dAwICOjUqVOdOnVIkhw7dqxCoYAYoz6rBP+1trZet26d+lLARyzLymSyDRs26LlKwkI9fvy4ffv2JEl+8sknnTt3btOmja2traur6969eyssT2CoGRkZderUgcW3s7MDtWRnZ2djYwP/PX78eIVfAb8aNGgQ7LK1tTW8yNLS0tbWFtZHT44W3nL69GkvLy+ZTAZKytfXVyqVNmnS5Nq1a7rHXyZFaGdnt3PnTg8Pj5CQkGfPnglfeP78+bBhwxBChw4dMpRk37FjB0Jo2rRpb9++FT66d+9eQECAm5vbq1evWJbVZ9XgLd27dx8+fDjP80VFRTzPX7p0ycLCYuzYseoT11Ogz58/v3HjxgqForCwsKioqLCwsLi4OD093dbW9vDhw3ouF6zA1atXpVJpz54979+/L3yUnp4+bdo0R0fH6OhoW1vb27dv668Ib968aWVlhRD69ddfhQlyHAdT+PTTT2mabteunUEU4cyZM1u2bCmqCKdMmaL5UYUnFR4eTlHUH3/8UVxcLHx08uRJDw8PPz+/vLw8PVUUrI9KperWrZu9vf3BgweFp6lUqi1btiCEpk6daigj0oCAiQ8cOFAqlW7ZskWhUAgfHTt2zM3N7auvvho0aNDgwYMNogjbtWv3xRdfKJXKvLy8IjXos/jw5MmTJyOENmzYAGwOOHHihKOjY/fu3UGSVPgtwg/d3NyioqJEFaGTk9PmzZsNskp37961sbEJDAwEjga8e/duwYIFCKFNmzbpwxccxxUXFxcVFeXl5SmVyhkzZrRo0UKhUBQUFMBe6C9y+/TpExoaWlhYmJ2dnZeXl5ub++bNm6VLlyKE4uPj9eQC+O2ePXsQQqNHj3758qXwUXJycnh4OEmSp0+f1rFEpSvCtLS06tWrW1lZzZw5U9gYhmGEJ86ZM0cikbx+/VowtCtsO4MTs3jxYtEX+fr6duvWTU85CEvWtWvXoUOHwr8PHDhAEATYuQYxz4Hov/3220aNGpX4qKCgwMrKSk+7ASzZ3NxcFxcX8Gs112rt2rWWlpZWVlYG8Qj//fdfZ2fnBQsWSCSSlJQU2GgY/7Jly9zd3SdOnFi/fn2DKEJgQlFFOHnyZM2PKkYAy5cvRwj9999/wh+Frc/NzXV3dx8yZIhBmHP+/PlSqfTFixeCahRedO7cOYTQ0aNHTUoXwkgWLVoE4SzN9UlPT/fx8bG1tR0/frxBRHzr1q0nT55sEBdcfQr79+9HCJ07d059CvDGlJQUiqKWLFmiz/iF0bq4uPz++++iitDBweGPP/7Q8y2gpTw9Pbt3767O7ALNrF69unbt2mA66LmG8PNZs2YBoxlwO3r27BkeHq65y35+fnoSEuiOp0+fIoS+++47UXk4efJkGxubjIwMbUtUuiLMzMyUSqUNGzaEsapzLBi8PM/XqVMHSLnCkxHCYnXr1gVHTaFQqN6jsLCQYZjLly8LwqvColBQhBERETzPr1u3DiG0ceNGGINBWFFdEQpRJvjvu3fv9FeE8Pxly5ZZWFgUFBRwHKe+7CzLQrC3V69ekHLTXxHeuHGDoqjs7Ozu3bt36dJFWKuXL18ihOLi4latWuXh4WEoRdi8eXOGYZRKJfMexcXFKpVq0qRJeipCGF5+fr6FhQXYW8XFxepjBu/n/PnzCCHwsyv2LnhmXl6eTCYDN7qgoEClBoi+9u/f3yA+rgGDomAKSKXSZcuWaVufCxcuIIQmTpxoKEU4adIkiCFzatDzsQ0bNhw2bBhst/rK5+fng6a3tbXVR3moK8KNGzeqVKri4mIgV5CTSqVSf0UIP9yyZQtJkpmZmSUeJURlQMTrufXCFsycObN58+bwcP2Vq6AIw8LCgJdhlYCWOnToAKG4Ci8RPD8yMrJBgwbwHHVuEiwGNze3r7/+WtuLSi88IQhCqVSOGDFCs6ZZ+PeYMWP+/vtvSHFXLM8Jj4qLi5s0aRJCSC6XS6VS+j0sLCwoimrbtq2np+ehQ4eE6m19sugWFhYbNmyYNGnSwYMHx40bp1KpaJo2eOUCoQH9nwkp3927dwcHB1taWrIsq77sUA7HcdyYMWMMOAvIqP/555+nTp3as2cPrNXIkSP9/f179er19u3bCm+95tZIJBKKouC/AJlMBmTAl7nIWVvFAULowoULCoVi5MiRHMdJJBL1TZFKpRzHdejQwcPDIzY2tsKUBr+6cuWKEKCztLSk1WBpaYkQmjlz5u3bt7OyskiS1HNqhqpLBC+KZdkRI0ZoW5/27dtXr149Pz/fBDkFaiJev3798OHDOXPmIISsrKzUVx6C/JMmTSosLExISIDyED0p1tbWlqZpmUwG5ErTNBCw/g+Hddi1a1fv3r0dHBxKFMEKZwMcHR31lC2aK29YqcXzvFQqBV6GVZJKpRcvXrx8+XJYWBiq6FEcnucpimIY5siRI59//jmstvqjKIqCPw4fPjwmJgZpOZZDl4WwaJpu1KiRaHkY/KVp06bZ2dn5+fkVK1CEnxQVFRUXF+/YsePChQssy5Z4CM/zNE0nJyc/ePBA/11xcHDYuXNnVFRUfHx8QECAQqGQyWRmUW3Mv68DfvnyZWRkpCibwUkpb29vVNrxoHJp37y8PG9v7x9//HHkyJHBwcEHDx48ceLEkydPoKrbUBO0srK6f/9+7969FQqFetWoVCq9e/eul5eXQQpTnZycXF1dReUUWIj169d/+PChni96+fIlQRADBw4Em1qz2jYrK0upVD5//tzBwYE3gQpSGPb9+/ddXV2dnJx0lNfVr19fqVSaLJu8evWK5/np06fL5XLRlQd359GjR1CErE+loo2NzcKFC7du3apUKtVfRJJkbm6unoIFmP358+fBwcE6xsmbavmxADs7u8OHD7dv3x74GuRJYmLi/Pnzu3TpInrKpey6IyMjIzc3t0mTJqJqG9i8WbNmmzZtghdpLlfp7xbOz2rbBv79kVs9dwK2nKZpa2trKOTV/MLo0aO7deuG9D7Jm5+fHxAQwDDMggULDhw4AH6VeZ3FgVMfulnUsG+USCQ8z8+dO3fHjh2RkZHnz5//4Ycf6tWrhwx6yJ3jOJlM5unpyTCMQANgjb148ULPSIBgJGpqJs211Z8ewBrVcQbAxsamSZMmcDDJdAQZSZJlWR9T5g7wsGUymZWVlaihRpLk559/7uvrq+fKg051cXEBilVnBIIgoKLeINNhGEZ35b+JK0KVSuXq6tq5c2eVSiXEP3x9fbdu3ern59evX7+KHY2DWcMPS5xVK4ESu1Nuj5AgCIZhbt682bdvX1HzmSCIK1euuLq6WllZVWw/QGPLZLJq1ar17dt33rx5ZdSaFYZCoahXr96aNWtatWrVtm3bCxcuVKtWzbC6EGKtBjfWBBldr169S5cuzZw5U1SXEARx584dZLj2HOrRku3bt7do0aJ58+ZfffWVUqmUSqUGfEVRUZGnp+fatWs1P5o7d+4///yjf6CpadOmGRkZKSkpHh4eSOxYJBTp9e/fX88XeXp6siy7ZcuWsrgFpiDIYAzNmjV78+bNq1ev3N3dS/AaSACO4xITE/X0zit1CrVr16Zpev78+S1btqxUYUKSZEFBwcSJE0NDQzU/3bt3r55+MyQ+GjRo8O+//+qwS0y/OUB+fn6bNm2+//77En9fsGBBZGRkamqqpaVlBUQlfN/JycnR0fHff//t0aOHNiV1+fLlmjVr0jQtqnFLpwCO4+Ry+aZNm5RKJRgm8CbIQ0J8dtOmTZGRkfp4IfDD0NDQ1atXK5VKpVIp5FQh7axQKAzo6JAkmZ+fL5PJrl275uDg0Lx58zdv3sBcDOLT8Dyfmprq4OAgyA6Dx69GjRoVGxublpZG0zSkf9XXhyCINWvWVJJl17x586+//nr58uVC/N2wgkyoNVAvlmEYJj8/X08DCEzRtm3bOjo6rlq1iiRJqA4QFhaIPCYmJj09HU4+VeyN8CI41vnjjz8ihKBYRpgRVAoI5rxJOVIBAQF2dnarV68Gfi+xPhRFxcbGZmVlWVtbm6Yi5DjOycmpVatW8+fPh5VXpyWowlNnJf3fmJ2drV4GIhCw/lkD+PmIESNOnz797NkzmqahFE4QNVAYcvv2bYPIrkolraKiIqh8FCS8UqkcP368QqF4+PBhhfOp4OqFhYX99ttvoORKKCmapouKirZu3QqH/USjSqUzORQBI4QiIiJIkoQqCdCx8O/IyEiKombMmAF5S32CSNOmTWMYZuLEiVKpVC6XQ86ZpmmpVCqTyWbNmjVkyBCI2xgq0EdR1NmzZ5s1a9a0adOnT5/SNK0/PUmlUoIgzpw50759e01m0992A6Nm6NChDRs2HDBgAMMw6ml5WLEZM2YkJSXZ2toaPJEDfX9+/PHHzp0767PjupmfFoP+AViQklKpdPXq1atXrz5y5IhcLofHAknLZLKnT59GRkZ++eWXtWrVgl5c+rxo6dKlP/zww/Xr162srITyHyisyM7ObtWq1cmTJ00n0gjDlsvlv/zyy/Lly48fPy6VSkusz507d6ZPn+7q6mqyOUJghOXLlx89enT79u1WVlZQeSesvEQi6dGjx6+//mqQlYeKFVGK1Z/ZIYz/2Wef+fv79+3bt7i4GMSLUBVCUdR3333Xrl27vLw8kzKqREWuRCKxtLSUSqUg4aVSaVJSEpTX6rNEPM/PmzevsLBw5MiRmkoKIRQSEuLk5DRhwgTRZodlUoQkSb579+73339//vx5mzZtLly4AAk8lUoFlSZHjhw5evSora2tPjFAGLe9vf2RI0e2b9/eq1evGzduCCUz9+/fj4iIWLFiRd++fQ3ILcD2HMfFxsb27du3adOmN2/eBDVT4ccWFBQ8e/ZswoQJqampY8aM4TV63BlEiwNVHT58OCUlxdfXNy4urri4GKaTkJDQp0+f/fv379mzRzTVahBdqDuHZBBBVt6PyjX4iIiIuXPn9u3bd+7cuSkpKUCBmZmZ69evb9y4cZ8+fX766acKa0F1i2Hs2LETJkxo3br1zz//nJaWBlGswsLCmJiYxo0bq1QqX19fk2rWCsMeOXLkjBkzevTo8fXXX7948QLWJyMjA7IJX3/9dUBAQE5OTmVvtz5b7O/v/+uvvw4dOnTq1KnJyclCsOHChQstW7a8efPmp59+yhuoy1oFPiovYmNjFQpFw4YNDxw4UFBQIAjGyMjIH374Ydu2bfb29qacKaQoKi8v7/Xr18nJySkpKSkpKUlJSUeOHAkKCurVq1edOnUqzAWgO5ydnY8ePbpnz56OHTteunQJRJ9SqTx37lzr1q2vXLkSFxdnYWGh1RUpS/cdiUQSHx+vUCgiIiIsLS3d3Nzq1avn4uJCUVSPHj2ePn1qqINQ8JDExMROnTpJpVJ3d3cfHx8PDw+ZTObn5wfF6AZpsdatWzc4YwTn4SC0OGvWLDs7u3PnzlWsMwA85M8//5TL5R4eHnFxceqjFU5oWVtbG6QRDzw5LS0tLCxMIpE4Ojp6eXm5u7tbWFj06dMnMzPz6dOnBEHcvXtX/3OE169fpygqKSlJ81FCG5169eoZ6hyhn5+ftgP1mh/pQwbQL0kul9euXbtevXo2NjYODg5wzlrPNnsljnlt2LChevXq1tbWXl5ePj4+jo6OVlZWkydPhqY2JnKIUHPYW7ZsqVGjhlQqhfWxtbWtWbPmX3/9xfN8ly5d4Hy0/ucI27Rpo+cpZB3kFBcX16BBAwsLizp16vj4+Li5ucnl8oEDB+rZv41Xa7Hm6uqq7UC9k5OTnucI1R+Yk5MzZswYS0tLR0fHevXq1apVSy6Xt2zZ8urVq4aiIqG7k/6dK0rwWnh4OELI2tpa+h4ymUwul4eGhmZnZxvqtOKDBw+6du1K07SrqysoKalUOmDAAKgi1jEdotSS3MzMzBo1ahw5cgRaOb969So+Pj49Pd3e3r5t27ZQNGhAk1Z41JMnT/7999/MzEwbG5uWLVtCe19DNd1+/vw5TdNQCwAGBbhu9+7dk0gkXl5eFcg/C4W8aWlpPj4+kG4p8RCWZR8/fuzu7q6nA11irV6+fHnp0qW3b9/a29v7+/tD2+j8/PwXL154enrqfzKkoKAgKSnJ29tbJpNpHmuB1tXv3r0zyNmGN2/eFBUV1a1bV7OGBT6qU6eOQcxeWD2O465cuXL79m2WZb29vQMCAipc86X7RSqV6vLly4mJiQzD1K5du127do6OjsiEC99h2CzLXrx4Ecyphg0bdujQAUImL168oCjK3d1dz0oNnuefPXsml8vd3NyQoSuGhPK3Gzdu3Lp1q7Cw0NXVtW3btlAkZRCpxfP848ePnZ2dNR0y+MjFxcXOzs6AzJ6RkXHlypVXr15BN2Y42GaoQj/gwbS0tMLCwhI8qOczX7x4kZ2dXaKNfvXq1Z2dnQ1LsQihp0+fXrlyJTs729nZuV27drVq1Sp1u8uqCI8ePdqxY0eVSlWiRLAyrojTdlkgX5nXMCFDXydknHgXmFEleABCymZ6NZIxYbRrhoxz35P5ro+Rp6DtPiMTBy92nxev5ZIvc4Fh90JUd5RFHpbpDCOMEnLCggMLo68MaQvPFF6E3p8QN+wVcZrmJ/gH+lu4sDjaVsawGy8UMghJO/WLKCvDU9cxa4O8S8ejDG4MQZpdOCMLu1YZMsVoLzLasEU5yBREoeYUkNpltpUhtXSM37BTA9YuweyVQUiV4XWIFvIYfC9K6I4S8lBfRaheWGU0P6NSX6SNNA1Sl6ib7itjXto224DvKuVaS4NyuzHPS5WRT8zoRcYZtgE3wggixeC30pdx/MZkdsO+opKaTZqs7ijTDywsLMwujICBgYGBgVEmPV1q7TLDMCkpKW5ublgdYmBgYGB8jIoQAwMDAwPjA0aZQqMGPBaKgYGBgYGBPUIMDAwMDAyz8ggxMDAwMDCwIsTAwMDAwMCKEAMDAwMDAytCDAwMDAwMrAgxMDAwMDA+BPw/U+WLHbIOrs8AAAAASUVORK5CYII="

doc_storage = {}
doc_counter = 0

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp  = Dispatcher()

CITY_REGION = {
    'москва': 'Москва и Московская область',
    'воронеж': 'Воронеж, Воронежская область',
    'ярославль': 'Ярославль, Ярославская область',
    'тула': 'Тула, Тульская область',
    'рязань': 'Рязань, Рязанская область',
    'липецк': 'Липецк, Липецкая область',
    'санкт-петербург': 'Санкт-Петербург и Ленинградская область',
    'калининград': 'Калининград, Калининградская область',
    'мурманск': 'Мурманск, Мурманская область',
    'архангельск': 'Архангельск, Архангельская область',
    'вологда': 'Вологда, Вологодская область',
    'петрозаводск': 'Петрозаводск, Республика Карелия',
    'ростов-на-дону': 'Ростов-на-Дону, Ростовская область',
    'краснодар': 'Краснодар, Краснодарский край',
    'волгоград': 'Волгоград, Волгоградская область',
    'сочи': 'Сочи, Краснодарский край',
    'астрахань': 'Астрахань, Астраханская область',
    'симферополь': 'Симферополь, Республика Крым',
    'ставрополь': 'Ставрополь, Ставропольский край',
    'махачкала': 'Махачкала, Республика Дагестан',
    'владикавказ': 'Владикавказ, Республика Северная Осетия',
    'грозный': 'Грозный, Чеченская Республика',
    'нальчик': 'Нальчик, Кабардино-Балкарская Республика',
    'пятигорск': 'Пятигорск, Ставропольский край',
    'казань': 'Казань, Республика Татарстан',
    'нижний новгород': 'Нижний Новгород, Нижегородская область',
    'уфа': 'Уфа, Республика Башкортостан',
    'самара': 'Самара, Самарская область',
    'пермь': 'Пермь, Пермский край',
    'саратов': 'Саратов, Саратовская область',
    'екатеринбург': 'Екатеринбург, Свердловская область',
    'челябинск': 'Челябинск, Челябинская область',
    'тюмень': 'Тюмень, Тюменская область',
    'сургут': 'Сургут, Ханты-Мансийский АО',
    'магнитогорск': 'Магнитогорск, Челябинская область',
    'нижний тагил': 'Нижний Тагил, Свердловская область',
    'новосибирск': 'Новосибирск, Новосибирская область',
    'красноярск': 'Красноярск, Красноярский край',
    'иркутск': 'Иркутск, Иркутская область',
    'кемерово': 'Кемерово, Кемеровская область',
    'барнаул': 'Барнаул, Алтайский край',
    'томск': 'Томск, Томская область',
    'владивосток': 'Владивосток, Приморский край',
    'хабаровск': 'Хабаровск, Хабаровский край',
    'благовещенск': 'Благовещенск, Амурская область',
    'якутск': 'Якутск, Республика Саха (Якутия)',
    'улан-удэ': 'Улан-Удэ, Республика Бурятия',
    'магадан': 'Магадан, Магаданская область',
}

def next_month_range():
    today = date.today()
    year  = today.year + (1 if today.month == 12 else 0)
    month = 1 if today.month == 12 else today.month + 1
    from calendar import monthrange as mr
    days  = mr(year, month)[1]
    start = date(year, month, 1)
    end   = date(year, month, days)
    return f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}", days

def days_word(n):
    if n == 31: return '31 день'
    if n == 30: return '30 дней'
    return '28 дней'

def parse_budget(s):
    import re
    nums = re.findall(r'\d+', str(s).replace(' ','').replace('\u202f','').replace('\xa0',''))
    return max((int(n) for n in nums), default=500000)

def create_excel_bytes(data):
    city     = data.get('city', '')
    brand    = data.get('brand', '')
    budget   = parse_budget(data.get('budget', '500000'))
    calls    = int(data.get('calls') or 50)
    channels = data.get('channels', ['Яндекс Директ'])
    region   = CITY_REGION.get(city.lower(), city)
    date_range, n_days = next_month_range()
    period   = days_word(n_days)

    GREEN  = PatternFill('solid', fgColor='92D050')
    LIGHT  = PatternFill('solid', fgColor='E2EFDA')
    thin   = Side(style='thin')
    def brd(): return Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Контекст'

    # Ширина A=134px, B=123px (формула: width ≈ (px - 5) / 7 для шрифта по умолчанию)
    ws.column_dimensions['A'].width = (134 - 5) / 7   # ≈ 18.43
    ws.column_dimensions['B'].width = (123 - 5) / 7   # ≈ 16.86
    # Высота строки 1 — увеличена до 126 px чтобы вместить логотип высотой 1.31"
    # (формула: pt = px * 72/96, при 96 DPI)
    ws.row_dimensions[1].height = 126 * 72 / 96       # = 94.5 pt = 126 px

    # Логотип на A1:B1
    ws.merge_cells('A1:B1')
    logo_bytes = base64.b64decode(LOGO_B64)
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    tmp.write(logo_bytes)
    tmp.close()
    xl_img = XLImage(tmp.name)
    # Размеры по ТЗ: 3.64" × 1.31" (при 96 DPI = 349×126 px)
    xl_img.width  = 349   # 3.64 дюйма
    xl_img.height = 126   # 1.31 дюйма
    xl_img.anchor = 'A1'
    ws.add_image(xl_img)

    # Мета
    for row, label, value in [
        (2, 'Посадочная страница', 'ваш ленд'),
        (3, 'Срок кампании',       date_range),
        (4, 'Регион показа',       region),
    ]:
        ws[f'A{row}'].value     = label
        ws[f'A{row}'].font      = Font(name='Arial', size=8, bold=True)
        ws[f'A{row}'].alignment = left_a
        ws[f'B{row}'].value     = value
        ws[f'B{row}'].font      = Font(name='Arial', size=8)
        ws.merge_cells(f'B{row}:M{row}')

    # Заголовки
    headers = [
        ('A','B','Поисковая система / Канал'),
        ('C','C','Рекламный носитель'),
        ('D','D','Время размещения'),
        ('E','E','Охват (прогноз)'),
        ('F','F','Показы (прогноз)'),
        ('G','G','Средний CTR,%'),
        ('H','H','Переходы'),
        ('I','J','Средняя цена за клик'),
        ('K','K','РЛ'),
        ('L','L','Стоимость РЛ'),
        ('M','M','Бюджет с НДС и АК'),
    ]
    for sc, ec, title in headers:
        if sc != ec: ws.merge_cells(f'{sc}5:{ec}5')
        c = ws[f'{sc}5']
        c.value = title
        c.font  = Font(name='Arial', size=8, bold=True)
        c.fill  = GREEN
        c.alignment = center
        c.border = brd()
    ws.row_dimensions[5].height = 32

    # Данные
    dr = 6
    for ch in channels:
        is_ya = 'яндекс' in ch.lower()
        is_vk = 'вконтакте' in ch.lower() or 'vk' in ch.lower()
        ch_budget   = min(budget, 700000) if is_vk else budget
        cpc         = round(random.uniform(65, 80))
        clicks      = round(ch_budget / cpc)
        ctr         = round(random.uniform(0.0075, 0.011), 5)
        impressions = round(clicks / ctr)
        sys_name  = 'Яндекс.Директ' if is_ya else ('ВКонтакте' if is_vk else ch)
        placement = 'сети/поиск'     if is_ya else ('лента/сторис' if is_vk else '')

        ws.merge_cells(f'A{dr}:B{dr}')
        ws[f'A{dr}'].value = sys_name
        ws[f'A{dr}'].font  = Font(name='Arial', size=10, bold=True)

        for col, val in [('C', placement), ('D', period)]:
            ws[f'{col}{dr}'].value     = val
            ws[f'{col}{dr}'].font      = Font(name='Arial', size=10)
            ws[f'{col}{dr}'].alignment = center

        for col, val, fmt in [
            ('E', impressions, '#,##0'),
            ('F', impressions, '#,##0'),
            ('H', clicks,      '#,##0'),
            ('K', calls,       '#,##0'),
            ('M', ch_budget,   '#,##0 ₽'),
        ]:
            ws[f'{col}{dr}'].value         = val
            ws[f'{col}{dr}'].font          = Font(name='Arial', size=10)
            ws[f'{col}{dr}'].alignment     = center
            ws[f'{col}{dr}'].number_format = fmt

        ws[f'G{dr}'] = f'=H{dr}/F{dr}'
        ws[f'G{dr}'].number_format = '0.00%'; ws[f'G{dr}'].alignment = center; ws[f'G{dr}'].font = Font(name='Arial', size=10)
        ws.merge_cells(f'I{dr}:J{dr}')
        ws[f'I{dr}'] = f'=M{dr}/H{dr}'
        ws[f'I{dr}'].number_format = '#,##0 ₽'; ws[f'I{dr}'].alignment = center; ws[f'I{dr}'].font = Font(name='Arial', size=10)
        ws[f'L{dr}'] = f'=M{dr}/K{dr}'
        ws[f'L{dr}'].number_format = '#,##0 ₽'; ws[f'L{dr}'].alignment = center; ws[f'L{dr}'].font = Font(name='Arial', size=10)
        dr += 1

    # Итог
    first, last = 6, dr - 1
    ir = dr
    ws.merge_cells(f'A{ir}:D{ir}')
    ws[f'A{ir}'].value = 'Итог:'; ws[f'A{ir}'].font = Font(name='Arial', size=10, bold=True); ws[f'A{ir}'].fill = LIGHT

    for col, fmt in [('E','#,##0'),('F','#,##0'),('H','#,##0'),('K','#,##0'),('M','#,##0 ₽')]:
        ws[f'{col}{ir}'] = f'=SUM({col}{first}:{col}{last})'
        ws[f'{col}{ir}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'{col}{ir}'].fill = LIGHT; ws[f'{col}{ir}'].alignment = center
        ws[f'{col}{ir}'].number_format = fmt

    ws[f'G{ir}'] = f'=H{ir}/F{ir}'
    ws[f'G{ir}'].font=Font(name='Arial',size=10,bold=True); ws[f'G{ir}'].fill=LIGHT
    ws[f'G{ir}'].number_format='0.00%'; ws[f'G{ir}'].alignment=center

    ws.merge_cells(f'I{ir}:J{ir}')
    ws[f'I{ir}'] = f'=M{ir}/H{ir}'
    ws[f'I{ir}'].font=Font(name='Arial',size=10,bold=True); ws[f'I{ir}'].fill=LIGHT
    ws[f'I{ir}'].number_format='#,##0 ₽'; ws[f'I{ir}'].alignment=center

    ws[f'L{ir}'] = f'=M{ir}/K{ir}'
    ws[f'L{ir}'].font=Font(name='Arial',size=10,bold=True); ws[f'L{ir}'].fill=LIGHT
    ws[f'L{ir}'].number_format='#,##0 ₽'; ws[f'L{ir}'].alignment=center

    for col, w in [('C',14),('D',12),('E',13),('F',13),('G',10),
                   ('H',10),('I',11),('J',11),('K',8),('L',14),('M',18)]:
        ws.column_dimensions[col].width = w

    out = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(out.name)
    with open(out.name, 'rb') as f:
        result = f.read()
    os.unlink(out.name)
    os.unlink(tmp.name)
    return result


@dp.message(CommandStart())
async def cmd_start(msg: Message):
    kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📋 Создать медиаплан", web_app=WebAppInfo(url=WEBAPP_URL))]],
        resize_keyboard=True,
    )
    await msg.answer(
        "👋 Добро пожаловать в <b>SINOBY MediaPlanBot</b>\n\nНажмите кнопку ниже чтобы создать медиаплан.",
        parse_mode="HTML", reply_markup=kb,
    )


@dp.message(F.web_app_data)
async def on_webapp_data(msg: Message):
    global doc_counter
    raw  = msg.web_app_data.data
    user = msg.from_user
    log.info(f"Получены данные от {user.id} (@{user.username}): {raw}")

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        await msg.answer("⚠️ Ошибка при получении данных.")
        return

    from datetime import datetime
    category    = data.get("category", "—")
    region      = data.get("region",   "—")
    district    = data.get("district", "—")
    city        = data.get("city",     "—")
    spec        = data.get("spec",     "—")
    goal        = data.get("goal",     "—")
    channels    = data.get("channels", [])
    budget      = data.get("budget",   "—")
    calls       = data.get("forecast_calls")
    price       = data.get("forecast_price")
    ts          = data.get("ts", "")

    spec_label  = "Бренд" if category == "Авто" else "Категория" if category == "Недвижимость" else "Направление"
    channel_str = ", ".join(channels) if channels else "—"

    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        time_str = dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        time_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    forecast_block = ""
    if calls and price:
        forecast_block = f"\n📊 <b>Прогноз</b>\n├ Звонков в месяц: <b>~{calls}</b>\n└ Цена звонка: <b>~{price}</b>\n"

    name     = user.full_name or "—"
    username = f"@{user.username}" if user.username else f"ID:{user.id}"

    text = (
        f"📋 <b>Новый запрос на медиаплан</b>\n{'─'*28}\n"
        f"🏷 Сфера: <b>{category}</b>\n"
        f"📍 Регион: <b>{region} ({district})</b>\n"
        f"🏙 Город: <b>{city}</b>\n"
        f"🔖 {spec_label}: <b>{spec}</b>\n"
        f"🎯 Цель: <b>{goal}</b>\n"
        f"📡 Канал: <b>{channel_str}</b>\n"
        f"💰 Бюджет: <b>{budget}</b>\n"
        f"{forecast_block}"
        f"{'─'*28}\n"
        f"👤 {name} · {username}\n🕐 {time_str}"
    )

    await msg.answer(
        f"✅ <b>Запрос принят!</b>\n\nНаш менеджер свяжется с вами в ближайшее время.\n\n"
        f"<b>Ваши параметры:</b>\n• Сфера: {category}\n• Город: {city}\n• {spec_label}: {spec}\n• Бюджет: {budget}",
        parse_mode="HTML", reply_markup=ReplyKeyboardRemove(),
    )

    doc_counter += 1
    doc_key = str(doc_counter)
    doc_storage[doc_key] = {
        "city": city, "brand": spec, "budget": budget,
        "calls": calls or "—", "price": price or "—",
        "channels": channels, "goal": goal, "category": category,
    }

    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="📄 Сформировать медиаплан", callback_data=f"doc:{doc_key}")
    ]])

    for manager_id in MANAGER_IDS:
        try:
            await bot.send_message(chat_id=manager_id, text=text, parse_mode="HTML", reply_markup=kb)
            log.info(f"Уведомление отправлено менеджеру {manager_id}")
        except Exception as e:
            log.error(f"Не удалось отправить менеджеру {manager_id}: {e}")


@dp.callback_query(F.data.startswith("doc:"))
async def on_create_doc(cb: CallbackQuery):
    await cb.answer("Генерирую файл...")
    await cb.message.edit_reply_markup(reply_markup=None)

    doc_key = cb.data[4:]
    data    = doc_storage.get(doc_key)

    if not data:
        await cb.message.answer("⚠️ Данные устарели. Попросите клиента пройти квиз заново.")
        return

    try:
        xlsx_bytes = create_excel_bytes(data)
        city  = data.get('city', '')
        brand = data.get('brand', '')
        today = date.today()
        year  = today.year + (1 if today.month == 12 else 0)
        month = 1 if today.month == 12 else today.month + 1
        fname = f"МП_{brand}_{city}_{month}.{year}.xlsx"

        await cb.message.answer_document(
            BufferedInputFile(xlsx_bytes, filename=fname),
            caption=f"📄 <b>Медиаплан готов</b>\n{brand} · {city}",
            parse_mode="HTML",
        )
    except Exception as e:
        log.error(f"Ошибка создания Excel: {e}")
        await cb.message.answer(f"⚠️ Ошибка: {e}")


@dp.message(Command("help"))
async def cmd_help(msg: Message):
    await msg.answer("ℹ️ <b>SINOBY MediaPlanBot</b>\n\n/start — квиз медиаплана\n/help — справка", parse_mode="HTML")


async def main():
    log.info("Бот запускается...")
    await dp.start_polling(bot, skip_updates=True)


if __name__ == "__main__":
    asyncio.run(main())
